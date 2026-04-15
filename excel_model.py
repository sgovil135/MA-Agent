"""
Excel Financial Model Builder for M&A Deal Bot.

Builds a 4-tab workbook:
  1. Financials — historical P&L + key metrics
  2. Valuation — EV/EBITDA multiples, purchase price bridge
  3. Deal Structure — 80/20 cash/note, debt capacity, DSCR
  4. Returns — Year 3 and Year 5 exit MOIC + IRR
"""

from __future__ import annotations

import io
from dataclasses import dataclass, field
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side, numbers
from openpyxl.utils import get_column_letter


# ── Data structures ───────────────────────────────────────────────────────

@dataclass
class FinancialData:
    """Extracted financials from a CIM."""
    company_name: str = "Unknown Company"
    sector: str = "mixed"  # "industrial", "defense_aerospace", "mixed"

    # Historical financials (up to 4 periods)
    periods: list[str] = field(default_factory=lambda: ["FY 2023", "FY 2024", "FY 2025", "LTM"])
    revenue: list[float | None] = field(default_factory=lambda: [None, None, None, None])
    gross_profit: list[float | None] = field(default_factory=lambda: [None, None, None, None])
    ebitda: list[float | None] = field(default_factory=lambda: [None, None, None, None])
    net_income: list[float | None] = field(default_factory=lambda: [None, None, None, None])
    capex: list[float | None] = field(default_factory=lambda: [None, None, None, None])
    owner_addbacks: list[float | None] = field(default_factory=lambda: [None, None, None, None])

    # Balance sheet items for purchase price bridge
    net_debt: float = 0.0

    # Key metrics
    customer_concentration: str = "Not disclosed"
    revenue_mix: str = "Not disclosed"
    employee_count: str = "Not disclosed"


# ── Sector multiples ──────────────────────────────────────────────────────

SECTOR_MULTIPLES = {
    "industrial": {"low": 4.0, "mid": 5.0, "high": 6.0, "label": "Industrial Services"},
    "defense_aerospace": {"low": 6.0, "mid": 7.0, "high": 8.0, "label": "Defense / Aerospace"},
    "mixed": {"low": 5.0, "mid": 6.0, "high": 7.0, "label": "Mixed / Other"},
}

# ── Styling constants ─────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=14)
SECTION_FONT = Font(name="Calibri", bold=True, size=11)
LABEL_FONT = Font(name="Calibri", size=11)
VALUE_FONT = Font(name="Calibri", size=11)
BOLD_VALUE_FONT = Font(name="Calibri", bold=True, size=11)
TITLE_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
TITLE_FONT = Font(name="Calibri", bold=True, size=14, color="FFFFFF")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
RESULT_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
THIN_BORDER = Border(
    bottom=Side(style="thin", color="999999"),
)
BOTTOM_BORDER = Border(
    bottom=Side(style="medium", color="333333"),
)

USD_FMT = '#,##0'
USD_K_FMT = '$#,##0'
USD_M_FMT = '$#,##0.0,,"M"'
PCT_FMT = '0.0%'
MULT_FMT = '0.0"x"'
PCT_DISPLAY_FMT = '0.0"%"'


def _set_col_widths(ws, widths: dict[int, float]):
    for col, w in widths.items():
        ws.column_dimensions[get_column_letter(col)].width = w


def _write_title(ws, row: int, col: int, text: str, span: int = 4):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = TITLE_FONT
    cell.fill = TITLE_FILL
    cell.alignment = Alignment(horizontal="left")
    for c in range(col, col + span):
        ws.cell(row=row, column=c).fill = TITLE_FILL


def _write_section(ws, row: int, col: int, text: str, span: int = 4):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    for c in range(col, col + span):
        ws.cell(row=row, column=c).fill = SECTION_FILL


def _write_label(ws, row: int, col: int, text: str, bold: bool = False):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD_VALUE_FONT if bold else LABEL_FONT


def _write_value(ws, row: int, col: int, value, fmt: str | None = None, bold: bool = False):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = BOLD_VALUE_FONT if bold else VALUE_FONT
    cell.alignment = Alignment(horizontal="right")
    if fmt:
        cell.number_format = fmt


def _write_result_row(ws, row: int, col_start: int, label: str, values: list, fmt: str | None = None, span: int = 4):
    """Write a highlighted result row."""
    _write_label(ws, row, col_start, label, bold=True)
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=col_start + 1 + i, value=v)
        cell.font = BOLD_VALUE_FONT
        cell.alignment = Alignment(horizontal="right")
        cell.fill = RESULT_FILL
        if fmt:
            cell.number_format = fmt
    for c in range(col_start, col_start + span):
        ws.cell(row=row, column=c).fill = RESULT_FILL


# ── Helpers ───────────────────────────────────────────────────────────────

def _safe(val: float | None, default: float = 0.0) -> float:
    return val if val is not None else default


def _adj_ebitda(data: FinancialData, idx: int) -> float:
    return _safe(data.ebitda[idx]) + _safe(data.owner_addbacks[idx])


def _ltm_adj_ebitda(data: FinancialData) -> float:
    """Get the most recent adj. EBITDA (last non-None period)."""
    for i in reversed(range(len(data.ebitda))):
        if data.ebitda[i] is not None:
            return _adj_ebitda(data, i)
    return 0.0


def _ltm_revenue(data: FinancialData) -> float:
    for i in reversed(range(len(data.revenue))):
        if data.revenue[i] is not None:
            return data.revenue[i]
    return 0.0


def _yoy_growth(current: float | None, prior: float | None) -> float | None:
    if current is None or prior is None or prior == 0:
        return None
    return (current - prior) / abs(prior)


def _gross_margin(gp: float | None, rev: float | None) -> float | None:
    if gp is None or rev is None or rev == 0:
        return None
    return gp / rev


def _ebitda_margin(ebitda: float | None, rev: float | None) -> float | None:
    if ebitda is None or rev is None or rev == 0:
        return None
    return ebitda / rev


def _seller_note_annual_payment(principal: float, rate: float = 0.06, years: int = 3) -> float:
    """Compute annual payment for a fully amortizing note (quarterly compounding)."""
    if principal <= 0:
        return 0.0
    n_periods = years * 4
    r = rate / 4
    if r == 0:
        return principal / years
    quarterly = principal * (r * (1 + r) ** n_periods) / ((1 + r) ** n_periods - 1)
    return quarterly * 4


def _senior_debt_annual_service(debt: float, rate: float = 0.07, years: int = 5) -> float:
    """Annual debt service on senior debt (assume 5yr amort, 7% rate)."""
    if debt <= 0:
        return 0.0
    n = years
    r = rate
    if r == 0:
        return debt / years
    annual = debt * (r * (1 + r) ** n) / ((1 + r) ** n - 1)
    return annual


def _remaining_debt(original: float, rate: float, years_total: int, years_elapsed: int) -> float:
    """Remaining balance on a fully amortizing loan."""
    if original <= 0 or years_total <= 0:
        return 0.0
    r = rate
    n = years_total
    if r == 0:
        return max(0.0, original * (1 - years_elapsed / n))
    annual_pmt = original * (r * (1 + r) ** n) / ((1 + r) ** n - 1)
    balance = original
    for _ in range(years_elapsed):
        interest = balance * r
        principal_pmt = annual_pmt - interest
        balance -= principal_pmt
    return max(0.0, balance)


def _irr(equity_invested: float, equity_proceeds: float, years: int) -> float | None:
    """Simple IRR for single cash-in / single cash-out."""
    if equity_invested <= 0 or equity_proceeds <= 0:
        return None
    return (equity_proceeds / equity_invested) ** (1.0 / years) - 1.0


# ── Tab builders ──────────────────────────────────────────────────────────

def _build_financials_tab(wb: Workbook, data: FinancialData):
    ws = wb.active
    ws.title = "Financials"
    _set_col_widths(ws, {1: 28, 2: 16, 3: 16, 4: 16, 5: 16})
    n_periods = len(data.periods)
    span = n_periods + 1

    r = 1
    _write_title(ws, r, 1, f"{data.company_name} — Financial Summary", span)

    r = 3
    _write_label(ws, r, 1, "", bold=True)
    for i, period in enumerate(data.periods):
        _write_label(ws, r, 2 + i, period, bold=True)
        ws.cell(row=r, column=2 + i).alignment = Alignment(horizontal="right")

    # Revenue
    r = 4
    _write_section(ws, r, 1, "Revenue", span)
    r = 5
    _write_label(ws, r, 1, "Revenue", bold=True)
    for i in range(n_periods):
        _write_value(ws, r, 2 + i, data.revenue[i], USD_K_FMT, bold=True)

    r = 6
    _write_label(ws, r, 1, "YoY Growth %")
    for i in range(n_periods):
        if i == 0:
            _write_value(ws, r, 2 + i, "—")
        else:
            g = _yoy_growth(data.revenue[i], data.revenue[i - 1])
            _write_value(ws, r, 2 + i, g, PCT_FMT)

    # Gross Profit
    r = 8
    _write_section(ws, r, 1, "Profitability", span)
    r = 9
    _write_label(ws, r, 1, "Gross Profit", bold=True)
    for i in range(n_periods):
        _write_value(ws, r, 2 + i, data.gross_profit[i], USD_K_FMT, bold=True)

    r = 10
    _write_label(ws, r, 1, "Gross Margin %")
    for i in range(n_periods):
        gm = _gross_margin(data.gross_profit[i], data.revenue[i])
        _write_value(ws, r, 2 + i, gm, PCT_FMT)

    # EBITDA
    r = 12
    _write_label(ws, r, 1, "EBITDA (Reported)")
    for i in range(n_periods):
        _write_value(ws, r, 2 + i, data.ebitda[i], USD_K_FMT)

    r = 13
    _write_label(ws, r, 1, "Owner Add-backs")
    for i in range(n_periods):
        _write_value(ws, r, 2 + i, data.owner_addbacks[i], USD_K_FMT)

    r = 14
    _write_label(ws, r, 1, "Adj. EBITDA", bold=True)
    for i in range(n_periods):
        adj = _adj_ebitda(data, i) if data.ebitda[i] is not None else None
        _write_value(ws, r, 2 + i, adj, USD_K_FMT, bold=True)

    r = 15
    _write_label(ws, r, 1, "EBITDA Margin %")
    for i in range(n_periods):
        adj = _adj_ebitda(data, i) if data.ebitda[i] is not None else None
        em = _ebitda_margin(adj, data.revenue[i])
        _write_value(ws, r, 2 + i, em, PCT_FMT)

    # Net Income + CapEx
    r = 17
    _write_label(ws, r, 1, "Net Income")
    for i in range(n_periods):
        _write_value(ws, r, 2 + i, data.net_income[i], USD_K_FMT)

    r = 18
    _write_label(ws, r, 1, "CapEx")
    for i in range(n_periods):
        _write_value(ws, r, 2 + i, data.capex[i], USD_K_FMT)

    # Key Metrics
    r = 20
    _write_section(ws, r, 1, "Key Metrics", span)
    r = 21
    _write_label(ws, r, 1, "Customer Concentration")
    _write_value(ws, r, 2, data.customer_concentration)
    r = 22
    _write_label(ws, r, 1, "Revenue Mix")
    _write_value(ws, r, 2, data.revenue_mix)
    r = 23
    _write_label(ws, r, 1, "Employee Count")
    _write_value(ws, r, 2, data.employee_count)


def _build_valuation_tab(wb: Workbook, data: FinancialData):
    ws = wb.create_sheet("Valuation")
    _set_col_widths(ws, {1: 32, 2: 16, 3: 16, 4: 16})

    multiples = SECTOR_MULTIPLES.get(data.sector, SECTOR_MULTIPLES["mixed"])
    adj_ebitda = _ltm_adj_ebitda(data)

    r = 1
    _write_title(ws, r, 1, "Valuation Analysis", 4)

    r = 3
    _write_label(ws, r, 1, "Adj. EBITDA (LTM)", bold=True)
    _write_value(ws, r, 2, adj_ebitda, USD_K_FMT, bold=True)

    r = 4
    _write_label(ws, r, 1, "Sector")
    _write_value(ws, r, 2, multiples["label"])

    # Column headers
    r = 6
    for i, label in enumerate(["Low", "Mid", "High"]):
        _write_label(ws, r, 2 + i, label, bold=True)
        ws.cell(row=r, column=2 + i).alignment = Alignment(horizontal="right")

    # Multiples
    r = 7
    _write_label(ws, r, 1, "EV/EBITDA Multiple")
    for i, key in enumerate(["low", "mid", "high"]):
        _write_value(ws, r, 2 + i, multiples[key], MULT_FMT)

    # Implied EV
    r = 8
    evs = [adj_ebitda * multiples[k] for k in ["low", "mid", "high"]]
    _write_result_row(ws, r, 1, "Implied Enterprise Value", evs, USD_K_FMT)

    # Purchase price bridge
    r = 10
    _write_section(ws, r, 1, "Purchase Price Bridge", 4)

    r = 11
    for i, label in enumerate(["Low", "Mid", "High"]):
        _write_label(ws, r, 2 + i, label, bold=True)
        ws.cell(row=r, column=2 + i).alignment = Alignment(horizontal="right")

    r = 12
    _write_label(ws, r, 1, "Enterprise Value")
    for i, ev in enumerate(evs):
        _write_value(ws, r, 2 + i, ev, USD_K_FMT)

    r = 13
    _write_label(ws, r, 1, "Less: Net Debt")
    for i in range(3):
        _write_value(ws, r, 2 + i, -abs(data.net_debt), USD_K_FMT)

    r = 14
    equity_values = [ev - abs(data.net_debt) for ev in evs]
    _write_result_row(ws, r, 1, "Equity Value", equity_values, USD_K_FMT)


def _build_deal_structure_tab(wb: Workbook, data: FinancialData):
    ws = wb.create_sheet("Deal Structure")
    _set_col_widths(ws, {1: 36, 2: 16, 3: 16, 4: 16})

    multiples = SECTOR_MULTIPLES.get(data.sector, SECTOR_MULTIPLES["mixed"])
    adj_ebitda = _ltm_adj_ebitda(data)
    evs = [adj_ebitda * multiples[k] for k in ["low", "mid", "high"]]
    equity_values = [ev - abs(data.net_debt) for ev in evs]

    r = 1
    _write_title(ws, r, 1, "Deal Structure — 80/20 Cash/Note", 4)

    # Column headers
    r = 3
    for i, label in enumerate(["Low", "Mid", "High"]):
        _write_label(ws, r, 2 + i, label, bold=True)
        ws.cell(row=r, column=2 + i).alignment = Alignment(horizontal="right")

    r = 4
    _write_label(ws, r, 1, "Total Consideration (Equity Value)", bold=True)
    for i, eq in enumerate(equity_values):
        _write_value(ws, r, 2 + i, eq, USD_K_FMT, bold=True)

    r = 5
    cash_at_close = [eq * 0.80 for eq in equity_values]
    _write_label(ws, r, 1, "Cash at Close (80%)")
    for i, c in enumerate(cash_at_close):
        _write_value(ws, r, 2 + i, c, USD_K_FMT)

    r = 6
    seller_notes = [eq * 0.20 for eq in equity_values]
    _write_label(ws, r, 1, "Seller Note (20%)")
    for i, s in enumerate(seller_notes):
        _write_value(ws, r, 2 + i, s, USD_K_FMT)

    # Seller note terms
    r = 8
    _write_section(ws, r, 1, "Seller Note Terms", 4)
    r = 9
    _write_label(ws, r, 1, "Term")
    _write_value(ws, r, 2, "3 years")
    r = 10
    _write_label(ws, r, 1, "Interest Rate")
    _write_value(ws, r, 2, 0.06, PCT_FMT)
    r = 11
    _write_label(ws, r, 1, "Payment Frequency")
    _write_value(ws, r, 2, "Quarterly")
    r = 12
    note_annual_mid = _seller_note_annual_payment(seller_notes[1])
    _write_label(ws, r, 1, "Annual Payment (P+I) — Mid Case")
    _write_value(ws, r, 2, note_annual_mid, USD_K_FMT)

    # Financing
    r = 14
    _write_section(ws, r, 1, "Financing", 4)

    r = 15
    for i, label in enumerate(["Low", "Mid", "High"]):
        _write_label(ws, r, 2 + i, label, bold=True)
        ws.cell(row=r, column=2 + i).alignment = Alignment(horizontal="right")

    r = 16
    senior_debt = adj_ebitda * 2.5
    _write_label(ws, r, 1, "Senior Debt Capacity (2.5x EBITDA)")
    for i in range(3):
        _write_value(ws, r, 2 + i, senior_debt, USD_K_FMT)

    r = 17
    equity_checks = [c - senior_debt for c in cash_at_close]
    _write_result_row(ws, r, 1, "Equity Check (Cash - Sr. Debt)", equity_checks, USD_K_FMT)

    # DSCR
    r = 19
    _write_section(ws, r, 1, "Debt Service Coverage — Year 1", 4)

    r = 20
    for i, label in enumerate(["Low", "Mid", "High"]):
        _write_label(ws, r, 2 + i, label, bold=True)
        ws.cell(row=r, column=2 + i).alignment = Alignment(horizontal="right")

    r = 21
    _write_label(ws, r, 1, "Adj. EBITDA")
    for i in range(3):
        _write_value(ws, r, 2 + i, adj_ebitda, USD_K_FMT)

    r = 22
    senior_annual = _senior_debt_annual_service(senior_debt)
    note_annuals = [_seller_note_annual_payment(sn) for sn in seller_notes]
    total_ds = [senior_annual + na for na in note_annuals]
    _write_label(ws, r, 1, "Total Debt Service (Sr. + Note)")
    for i, ds in enumerate(total_ds):
        _write_value(ws, r, 2 + i, ds, USD_K_FMT)

    r = 23
    dscrs = [adj_ebitda / ds if ds > 0 else None for ds in total_ds]
    _write_result_row(ws, r, 1, "Year 1 DSCR", dscrs, '0.00"x"')


def _build_returns_tab(wb: Workbook, data: FinancialData):
    ws = wb.create_sheet("Returns")
    _set_col_widths(ws, {1: 32, 2: 16, 3: 16, 4: 16})

    multiples = SECTOR_MULTIPLES.get(data.sector, SECTOR_MULTIPLES["mixed"])
    adj_ebitda = _ltm_adj_ebitda(data)
    evs = [adj_ebitda * multiples[k] for k in ["low", "mid", "high"]]
    equity_values = [ev - abs(data.net_debt) for ev in evs]
    cash_at_close = [eq * 0.80 for eq in equity_values]
    seller_notes = [eq * 0.20 for eq in equity_values]
    senior_debt = adj_ebitda * 2.5
    equity_checks = [c - senior_debt for c in cash_at_close]

    r = 1
    _write_title(ws, r, 1, "Returns Analysis", 4)

    r = 3
    _write_label(ws, r, 1, "Entry EBITDA", bold=True)
    _write_value(ws, r, 2, adj_ebitda, USD_K_FMT, bold=True)

    r = 4
    _write_label(ws, r, 1, "Assumption")
    _write_value(ws, r, 2, "EBITDA held flat (no growth)")

    for exit_year in [3, 5]:
        r_start = 6 if exit_year == 3 else 14

        r = r_start
        _write_section(ws, r, 1, f"Year {exit_year} Exit", 4)

        r = r_start + 1
        for i, label in enumerate(["Low", "Mid", "High"]):
            _write_label(ws, r, 2 + i, label, bold=True)
            ws.cell(row=r, column=2 + i).alignment = Alignment(horizontal="right")

        r = r_start + 2
        _write_label(ws, r, 1, "Exit Multiple (= Entry)")
        for i, key in enumerate(["low", "mid", "high"]):
            _write_value(ws, r, 2 + i, multiples[key], MULT_FMT)

        r = r_start + 3
        exit_evs = [adj_ebitda * multiples[k] for k in ["low", "mid", "high"]]
        _write_label(ws, r, 1, "Exit Enterprise Value")
        for i, eev in enumerate(exit_evs):
            _write_value(ws, r, 2 + i, eev, USD_K_FMT)

        r = r_start + 4
        remaining_sr = _remaining_debt(senior_debt, 0.07, 5, exit_year)
        # Seller note is 3yr, fully paid by year 3
        remaining_note = [
            _remaining_debt(sn, 0.06, 3, min(exit_year, 3)) for sn in seller_notes
        ]
        remaining_total = [remaining_sr + rn for rn in remaining_note]
        _write_label(ws, r, 1, "Less: Remaining Debt")
        for i, rd in enumerate(remaining_total):
            _write_value(ws, r, 2 + i, -rd, USD_K_FMT)

        r = r_start + 5
        equity_proceeds = [eev - rt for eev, rt in zip(exit_evs, remaining_total)]
        _write_result_row(ws, r, 1, "Equity Proceeds", equity_proceeds, USD_K_FMT)

        r = r_start + 6
        moics = [
            ep / ec if ec > 0 else None
            for ep, ec in zip(equity_proceeds, equity_checks)
        ]
        _write_result_row(ws, r, 1, "MOIC", moics, '0.00"x"')

        r = r_start + 7
        irrs = [
            _irr(ec, ep, exit_year) if ec and ec > 0 else None
            for ec, ep in zip(equity_checks, equity_proceeds)
        ]
        _write_result_row(ws, r, 1, "IRR", irrs, PCT_FMT)


# ── Public API ────────────────────────────────────────────────────────────

def build_financial_model(data: FinancialData) -> bytes:
    """Build the 4-tab Excel financial model. Returns .xlsx bytes."""
    wb = Workbook()
    _build_financials_tab(wb, data)
    _build_valuation_tab(wb, data)
    _build_deal_structure_tab(wb, data)
    _build_returns_tab(wb, data)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_financial_model_from_dict(extracted: dict) -> bytes:
    """Convenience wrapper that takes the dict from OpenAI extraction and builds the model."""
    data = FinancialData(
        company_name=extracted.get("company_name", "Unknown Company"),
        sector=extracted.get("sector", "mixed"),
        periods=extracted.get("periods", ["FY 2023", "FY 2024", "FY 2025", "LTM"]),
        revenue=extracted.get("revenue", [None, None, None, None]),
        gross_profit=extracted.get("gross_profit", [None, None, None, None]),
        ebitda=extracted.get("ebitda", [None, None, None, None]),
        net_income=extracted.get("net_income", [None, None, None, None]),
        capex=extracted.get("capex", [None, None, None, None]),
        owner_addbacks=extracted.get("owner_addbacks", [None, None, None, None]),
        net_debt=extracted.get("net_debt", 0.0),
        customer_concentration=extracted.get("customer_concentration", "Not disclosed"),
        revenue_mix=extracted.get("revenue_mix", "Not disclosed"),
        employee_count=extracted.get("employee_count", "Not disclosed"),
    )
    return build_financial_model(data)
